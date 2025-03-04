from django.shortcuts import redirect, render
from .models import *
from django.http import HttpRequest, HttpResponse, JsonResponse
import json
from django.contrib import messages
import openpyxl
from django.shortcuts import render
from django.http import HttpResponse
# Create your views here.
def packege(request):
    if request.user.is_authenticated:
        packeges = Packege.objects.all()
        data = { 'row' : packeges,
                'username':  request.session['username'] ,
                'img': request.session['img'],
                }
        return render(request , 'preset/index.html' , data) 
    else:
        return render(request , 'inspectors/login.html')


def add_new_packages(request):
    if request.user.is_authenticated:
        if  Packege.objects.filter(description = request.POST['packname']).exists(): 
            messages.success(request, "exist")
            return redirect('packege')
    
        else :
            packname = request.POST['packname']
            newPck = Packege()
            newPck.description = packname
            newPck.save()
            print('هذا الاي دي الجديد')
            print(newPck.id)
            messages.success(request, "success")
            return redirect('packege_details' , id = newPck.id)
    else :
            return render(request , 'inspectors/login.html')


# تفاصيل الباقة
def packege_details(request , id):
    if request.user.is_authenticated:
        row=[]
        #id packege
        print(id)
        pck =Packege.objects.get(id = id)
        titles = Term_title.objects.filter(packege_id = id)
        # استخراج جميع البنود التابعه للعنوان
        for t in titles :
            if Preset_terms.objects.filter(term_title_id = t.id).exists:
                pre = Preset_terms.objects.filter(term_title_id = t.id)
                row.append({'title_id': t.id ,'title': t.section , 'terms': pre})
        data = {
            'pckName' : pck.description ,
            'pkid': id ,
            'titles': titles ,
            'username':  request.session['username'] ,
            'img': request.session['img'],
            'row': row,
        }
        return render(request , 'preset/packages_details.html' , data) 
    else :
        return render(request , 'inspectors/login.html')

# رفع الملف    
def upload_packge_details(request , id):
  if request.method == "POST":
        excel_file = request.FILES.get("excel_file")
        if not excel_file:
            return HttpResponse("لم يتم اختيار ملف", status=400)
        try:
            wb = openpyxl.load_workbook(excel_file)
        except Exception as e:
            return HttpResponse("خطأ في قراءة الملف: " + str(e), status=400)
        
        # استخدام الورقة النشطة
        sheet = wb.active

        # التحقق من هيدر الملف (الصف الأول)
        header = [sheet.cell(row=1, column=i).value for i in range(1, 3)]
        if len(header) < 2 or not (header[0] and header[1] and 
           header[0].strip().lower() == "type" and header[1].strip().lower() == "text"):
            messages.error(request, "يوجد خطأ في هيكلة الملف (الهيدر غير صحيح)")
            return redirect("packege_details", id=id)
        
        # الحصول على باقي الصفوف بعد الهيدر وتخزينها في قائمة
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        if not rows:
            messages.error(request, "الملف فارغ بعد الهيدر")
            return redirect("packege_details", id=id)
        
        # التحقق من أن أول صف بعد الهيدر يحتوي على type = 0
        if rows[0][0] != 0:
            messages.error(request, "الصف الأول بعد الهيدر يجب أن يحتوي على type = 0")
            return redirect("packege_details", id=id)
        
        # التحقق من صحة الهيكلة لكل الصفوف
        for i, row in enumerate(rows):
            type_value = row[0]
            # التأكد من قبول فقط القيم 0 أو 1
            if type_value not in (0, 1):
                messages.error(request, f"في الصف {i+2} قيمة type يجب أن تكون 0 أو 1 فقط")
                return redirect("packege_details", id=id)
            # إذا كان الصف يحتوي على 0:
            if type_value == 0:
                # التأكد من وجود صف تالٍ يحتوي على 1
                if i+1 < len(rows):
                    next_type = rows[i+1][0]
                    if next_type != 1:
                        messages.error(request, f"الصف {i+2} يحتوي على 0، يجب أن يتبعه صف يحتوي على 1")
                        return redirect("packege_details", id=id)
                else:
                    messages.error(request, f"الصف {i+2} يحتوي على 0 ولا يوجد صف تالي يحتوي على 1")
                    return redirect("packege_details", id=id)
                # التحقق من عدم وجود صفين متتاليين يحتويان على 0
                if i+1 < len(rows) and rows[i+1][0] == 0:
                    messages.error(request, f"يوجد صفين متتاليين يحتويان على 0 في الصفوف {i+2} و{i+3}")
                    return redirect("packege_details", id=id)
        
        # بعد اجتياز التحقق من الهيكلة، نستطيع البدء بعملية الإدخال في قاعدة البيانات
        try:
            packege_instance = Packege.objects.get(id=id)
        except Packege.DoesNotExist:
            messages.error(request, "لا توجد باقة متاحة في النظام")
            return redirect("packege_details", id=id)
        
        current_title = None  # لتخزين العنوان الرئيسي الحالي
        results = []
        for row in rows:
            type_value = row[0]
            text_value = row[1]
            if type_value == 0:
                # إنشاء سجل للعنوان الرئيسي
                current_title = Term_title.objects.create(
                    section=text_value, 
                    packege_id=packege_instance
                )
                results.append(f"تم إنشاء عنوان: {text_value}")
                print(f"تم إنشاء عنوان: {text_value}")
            elif type_value == 1:
                if current_title:
                    # إنشاء سجل للبند وربطه بالعنوان الرئيسي الحالي
                    Preset_terms.objects.create(
                        description=text_value, 
                        term_title_id=current_title
                    )
                    results.append(f"تم إضافة بند '{text_value}' تحت العنوان '{current_title.section}'")
                    print(f"تم إضافة بند '{text_value}' تحت العنوان '{current_title.section}'")
                else:
                    results.append(f"بند '{text_value}' بدون عنوان رئيسي مرتبط")
                    print(f"بند '{text_value}' بدون عنوان رئيسي مرتبط")
        
        messages.success(request, "success_upload")
        return redirect("packege_details", id=id)
    

# صفحة اضافة تفاصيل تصنيف جديد
def add_termsPage(request , id):
    # id for Term_title
    if request.user.is_authenticated:
        title = Term_title.objects.get(id = id)
        terms = Preset_terms.objects.filter(term_title_id = id)

        data = {
            'titleID' : id ,
            'pkid' : title.packege_id_id,
            'pckName' : title.packege_id.description,
            'title': title.section,
            'username':  request.session['username'] ,
            'img': request.session['img'] ,
            'terms': terms,
        }
        return render(request , 'preset/new_terms.html' , data) 
    else :
        return render(request , 'inspectors/login.html')



# صفحة اضافة تفاصيل تصنيف جديد
def add_new_terms(request):
    if request.user.is_authenticated:
        terms = json.loads(request.POST['terms'])
        titleID = request.POST['titleID']
        for t in terms:
            new = Preset_terms()
            new.description = t
            new.term_title_id_id = titleID
            new.save()
        
        data = {
            'terms' : 'ok',
            # 'sec': request.POST['secName']
        }
        return JsonResponse( data , safe=False)
    else :
        return render(request , 'inspectors/login.html')




# تفاصيل التصنيف
def section_details(request , id ):
    if request.user.is_authenticated:
        #id title
        title = Term_title.objects.get(id = id).section
        terms = Preset_terms.objects.filter(term_title_id = id)
        data = {
            'title' : title ,
            'terms': terms,
            'pckName': Term_title.objects.get(id = id).packege_id.description,
            'pckid': Term_title.objects.get(id = id).packege_id.id,
            'username':  request.session['username'] ,
            'img': request.session['img']
        }
        return render(request , 'preset/section_details.html' , data) 
    else:
        return render(request , 'inspectors/login.html')


def add_new_term_title(request , id):
    #id pagege
    if Term_title.objects.filter(section = request.POST['description']).exists():
        messages.success(request, "exist")
        return redirect(packege_details , id =id )

    else :
        newTitle = Term_title()
        newTitle.section = request.POST['description']
        newTitle.packege_id_id = id
        newTitle.save()
        messages.success(request, "successadd")
        return redirect(packege_details , id =id )


def update_term_title(request , id):
    #id title
    update = Term_title.objects.get(id = id)
    pkid = update.packege_id_id
    if Term_title.objects.filter(section = request.POST['description']).exists():
        messages.success(request, "exist")
        return redirect(packege_details , id =pkid )
    else :
        update.section = request.POST['description']
        update.save()
        messages.success(request, "update")
        return redirect('packege_details' , id = pkid)
    
def deleteTitle(request , id):
    #id for title
    update = Term_title.objects.get(id = id)
    pkid = update.packege_id_id
    if Preset_terms.objects.filter(term_title_id = id).exists():
        messages.success(request, "exist_terms")
        return redirect('packege_details' , id = pkid)
    else:
        tt = Term_title.objects.get(id = id)
        tt.delete()
        messages.success(request, "delete")
        return redirect('packege_details' , id = pkid)


def delete_term(request , id):
    term = Preset_terms.objects.get(id = id)
    termid = term.term_title_id.id
    #packid = term.term_title_id.packege_id.id
    term.delete()
    messages.success(request, "delete")
    return redirect('add_termsPage' , id = termid)

def delete_package(request , id):
    # id for package
    pck = Packege.objects.get(id = id)
    if Term_title.objects.filter(packege_id = id).exists():
        return redirect(packege)
    else :
        pck.delete()
        return redirect(packege)
    
def edit_pck_title(request , id) :
    if  Packege.objects.filter(description = request.POST['description']).exists(): 
        messages.success(request, "exist")
        return redirect(packege)
    else :
        pck = Packege.objects.get(id = id)
        pck.description = request.POST['description']
        pck.save()
        messages.success(request, "update")
        return redirect(packege)
       